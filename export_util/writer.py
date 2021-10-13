import csv
from io import BytesIO, StringIO
from zipfile import ZipFile, ZIP_DEFLATED

import openpyxl
from openpyxl.writer.excel import ExcelWriter
from openpyxl import Workbook, load_workbook

## PDF
import itertools
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
##


class OutputTemplate:
    """
    This is the base class for the output writer
    template. You can extend this class to define
    your own template for the result document.
    """

    # Source template file path. This file will be used as the base file where
    # the formatted date will be written.
    template_file = None

    # Active worksheet index. You can select which worksheet will be used to
    # write formatted data.
    worksheet_index = 0

    # First cell from which table will be started to render.
    table_start = 'A1'

    # List of dicts which are describes images, which should be drawn to the sheet.
    # Use `image()` method to simplify listing process.
    images = []

    @staticmethod
    def image(cell_name, image_path, width=None, height=None):
        """
        Helper to simplify images listing for `OutputTemplate.images`
        :param cell_name:
        :param image_path:
        :param width:
        :param height:
        :return:
        """
        image = {
            'cell': cell_name,
            'name': image_path
        }

        if any([width, height]):
            image['size'] = {}

            if width is not None:
                image['size']['width'] = width
            if height is not None:
                image['size']['height'] = height

        return image


class BytesOutputWriter:
    """
    Simple data writer.

    Data writer `write_row` calls when we already have the formatted columns list.
    """
    mime_type = 'application/octet-stream'
    extension = 'bin'

    def __init__(self):
        self.output = BytesIO()

    def write(self, *cols):
        self.output.write(b''.join([x.encode() if isinstance(x, str) else x for x in cols]))

    def get_data(self):
        return self.output.getvalue()


class XLSXBytesOutputWriter(BytesOutputWriter):
    """
    Returns bytes of the XLSX workbook object.
    """
    mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    extension = 'xlsx'

    INTCOL_MAP = {i: c for i, c in enumerate(list('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))}
    COLINT_MAP = {c: i for i, c in enumerate(list('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))}

    def __init__(self, cols_dimensions=None, template=None):
        super(XLSXBytesOutputWriter, self).__init__()
        self.start_col = 0
        self.start_row = 0
        self._current_row = 0

        if template is not None:
            if template.template_file is not None:
                # Microsoft excel raises an error when using vba
                self.wb = load_workbook(filename=template.template_file, keep_vba=False)
            else:
                self.wb = Workbook()

            self.ws = self.wb.active
            self.from_template(template)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.resize_columns(cols_dimensions)

    def resize_columns(self, cols_dimensions):
        for col, dimension in (cols_dimensions or {}).items():
            if isinstance(dimension, dict):
                width = dimension.get('width')
                height = dimension.get('height')
            else:
                width = dimension
                height = None

            if width:
                self.ws.column_dimensions[col].width = width
            if height:
                self.ws.column_dimensions[col].height = height

    def from_template(self, template=None):
        if template.worksheet_index is not None:
            self.ws = self.wb.worksheets[template.worksheet_index]

        if template.table_start is not None:
            col, row = tuple(template.table_start)
            self.start_col = self.COLINT_MAP.get(col)
            self.start_row = int(row) - 1

        if template.images:
            for image in template.images:
                xlimg = openpyxl.drawing.image.Image(image['name'])
                if image['size']:
                    if image['size'][0]:
                        xlimg.width = image['size'][0]
                    if image['size'][1]:
                        xlimg.height = image['size'][1]
                self.ws.add_image(xlimg, image['cell'])

    def write(self, *cols):
        for i, val in enumerate(cols):
            if not val:
                continue

            try:
                self.ws.cell(row=self._current_row + self.start_row + 1, column=i+1, value=val)
            except Exception as e:
                print(e)

        self._current_row += 1

    def get_data(self):
        archive = ZipFile(self.output, 'w', ZIP_DEFLATED, allowZip64=True)
        writer = ExcelWriter(self.wb, archive)
        try:
            writer.write_data()
            writer.save()
        finally:
            archive.close()

        return self.output.getvalue()

    def _get_column_name(self, cell_name):
        return cell_name[0]


class CSVBytesOutputWriter(BytesOutputWriter):
    """
    CSV Data Writer.
    """
    mime_type = 'text/csv'
    extension = 'csv'

    def __init__(self, delimiter=';', template=None):
        super(CSVBytesOutputWriter, self).__init__()

        self.output = StringIO()
        self.writer = csv.writer(self.output, delimiter=delimiter)

        if template is not None:
            if template.template_file is not None:
                with open(template.template_file, 'r') as feed:
                    reader = csv.reader(feed, delimiter=delimiter)
                    for row in reader:
                        self.write(*row)
                feed.close()


    def write(self, *cols):
        self.writer.writerow([x for x in cols])

    def get_data(self):
        return self.output.getvalue()


class PDFBytesOutputWriter(BytesOutputWriter):
    """
    PDF Data Writer.
    """
    mime_type = 'text/pdf'
    extension = 'pdf'

    def grouper(self, iterable, n):
        args = [iter(iterable)] * n
        return itertools.zip_longest(*args)

    def __init__(self, offsets, pagesize, header=None):
        super(PDFBytesOutputWriter, self).__init__()

        self.output = BytesIO()
        self.writer = canvas.Canvas(self.output)

        self.offsets = offsets
        if header:
            self.data = header
        else:
            self.data = []

        self.w = pagesize[0]
        self.h = pagesize[1]

    def write(self, *cols):
        self.data.append(cols)


    def get_data(self):
        # Page size
        w = self.w
        h = self.h

        max_rows_per_page = 45
        # Margin.
        x_offset = 50
        y_offset = 50
        # Space between rows.
        padding = 15

        xlist = [x + x_offset for x in self.offsets]
        ylist = [h - y_offset - i * padding for i in range(max_rows_per_page + 1)]

        for rows in self.grouper(self.data, max_rows_per_page):
            rows = tuple(filter(bool, rows))
            self.writer.grid(xlist, ylist[:len(rows) + 1])
            for y, row in zip(ylist[:-1], rows):
                for x, cell in zip(xlist, row):
                    self.writer.drawString(x + 2, y - padding + 3, str(cell))
            self.writer.setPageSize((w, h))
            self.writer.showPage()

        self.writer.save()
        return self.output.getvalue()


__all__ = [
    'OutputTemplate',
    'BytesOutputWriter',
    'CSVBytesOutputWriter',
    'XLSXBytesOutputWriter',
    'PDFBytesOutputWriter'
]
